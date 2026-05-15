# 공용차량 예약 및 주행거리 정산 시스템
import streamlit as st
import psycopg
from psycopg.rows import dict_row
import streamlit.components.v1 as components
import pandas as pd
from datetime import datetime, date, timezone, timedelta

KST = timezone(timedelta(hours=9))

def now_kst():
    return datetime.now(KST)
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

_cal_component = components.declare_component(
    "yardchacha_calendar",
    path="./calendar_component",
)

st.set_page_config(
    page_title="공용차량 관리",
    page_icon="🚗",
    layout="wide",
    initial_sidebar_state="expanded",
)

DEPARTMENTS = [
    "생산기술팀", "공법기술", "의장기술", "운반종합기술",
    "시공기술", "족장기술", "DFX그룹", "도장기술",
]
ADMIN_PASSWORD  = "1111"
VEHICLE_NAME    = "EV3"
VEHICLE_NUMBER  = "05하 7211"
WEEKDAYS        = ["월", "화", "수", "목", "금", "토", "일"]


# ════════════════════════════════════════════════════════════════
# DB  (Supabase / PostgreSQL)
# ════════════════════════════════════════════════════════════════
def _get_conn():
    db = st.secrets["supabase"]
    return psycopg.connect(
        host=db["host"],
        port=int(db.get("port", 6543)),
        dbname=db.get("dbname", "postgres"),
        user=db["user"],
        password=db["password"],
        sslmode="require",
        connect_timeout=15,
        row_factory=dict_row,
    )


def init_db():
    with _get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                phone       TEXT PRIMARY KEY,
                employee_id TEXT NOT NULL,
                department  TEXT NOT NULL,
                name        TEXT NOT NULL
            )""")
            cur.execute("""
            CREATE TABLE IF NOT EXISTS reservations (
                id           SERIAL PRIMARY KEY,
                user_phone   TEXT NOT NULL,
                department   TEXT NOT NULL,
                name         TEXT NOT NULL,
                phone        TEXT NOT NULL,
                res_date     TEXT NOT NULL,
                res_time     TEXT NOT NULL,
                res_time_end TEXT NOT NULL DEFAULT '',
                destination  TEXT NOT NULL,
                purpose      TEXT DEFAULT '',
                created_at   TEXT DEFAULT to_char(NOW() AT TIME ZONE 'Asia/Seoul',
                                                  'YYYY-MM-DD HH24:MI:SS')
            )""")
            cur.execute("""
            CREATE TABLE IF NOT EXISTS driving_logs (
                id               SERIAL PRIMARY KEY,
                user_phone       TEXT NOT NULL,
                department       TEXT NOT NULL,
                name             TEXT NOT NULL,
                phone            TEXT NOT NULL,
                drive_date       TEXT NOT NULL,
                odometer_start   REAL,
                odometer_end     REAL,
                companions       TEXT,
                destination      TEXT,
                charging_amount  REAL DEFAULT 0,
                parking_location TEXT,
                depart_time      TEXT DEFAULT '',
                arrive_time      TEXT DEFAULT '',
                purpose          TEXT DEFAULT '',
                status           TEXT DEFAULT 'pre',
                created_at       TEXT DEFAULT to_char(NOW() AT TIME ZONE 'Asia/Seoul',
                                                      'YYYY-MM-DD HH24:MI:SS')
            )""")
            for ddl in [
                "ALTER TABLE reservations ADD COLUMN IF NOT EXISTS res_time_end TEXT NOT NULL DEFAULT ''",
                "ALTER TABLE reservations ADD COLUMN IF NOT EXISTS purpose TEXT DEFAULT ''",
                "ALTER TABLE driving_logs ADD COLUMN IF NOT EXISTS charging_amount REAL DEFAULT 0",
                "ALTER TABLE driving_logs ADD COLUMN IF NOT EXISTS depart_time TEXT DEFAULT ''",
                "ALTER TABLE driving_logs ADD COLUMN IF NOT EXISTS arrive_time TEXT DEFAULT ''",
                "ALTER TABLE driving_logs ADD COLUMN IF NOT EXISTS purpose TEXT DEFAULT ''",
            ]:
                try:
                    cur.execute(ddl)
                except Exception:
                    conn.rollback()
        conn.commit()


def _query(sql, params=(), one=False):
    with _get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params or None)
            rows = cur.fetchall()   # dict_row → list[dict]
    return (rows[0] if rows else None) if one else rows


def _exec(sql, params=()):
    with _get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, params or None)
        conn.commit()


# ── 사용자 ──────────────────────────────────────────────────────
def get_user(phone):
    return _query("SELECT * FROM users WHERE phone=%s", (phone,), one=True)

def register_user(phone, emp_id, dept, name):
    _exec("""
        INSERT INTO users (phone, employee_id, department, name) VALUES (%s,%s,%s,%s)
        ON CONFLICT (phone) DO UPDATE
        SET employee_id=EXCLUDED.employee_id,
            department=EXCLUDED.department,
            name=EXCLUDED.name
    """, (phone, emp_id, dept, name))

def auth_user(phone, emp_id):
    return _query(
        "SELECT 1 FROM users WHERE phone=%s AND employee_id=%s",
        (phone, emp_id), one=True,
    ) is not None


# ── 예약 ────────────────────────────────────────────────────────
def get_all_reservations():
    return _query("SELECT * FROM reservations ORDER BY res_date, res_time")

def get_reservations_by_date(d):
    return _query(
        "SELECT * FROM reservations WHERE res_date=%s ORDER BY res_time", (d,))

def check_reservation_conflict(res_date, start_time, end_time, exclude_id=None):
    """시간이 겹치는 예약 목록 반환 (exclude_id: 수정 시 자기 자신 제외)"""
    sql = ("SELECT * FROM reservations "
           "WHERE res_date=%s AND res_time < %s AND res_time_end > %s")
    params = [res_date, end_time, start_time]
    if exclude_id:
        sql += " AND id != %s"
        params.append(exclude_id)
    return _query(sql, params)

def add_reservation(user_phone, dept, name, phone,
                    res_date, res_time, res_time_end, dest, purpose=""):
    _exec(
        "INSERT INTO reservations "
        "(user_phone,department,name,phone,res_date,res_time,res_time_end,destination,purpose) "
        "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (user_phone, dept, name, phone, res_date, res_time, res_time_end, dest, purpose),
    )

def update_reservation(rid, dept, name, res_date, res_time, res_time_end, dest, purpose=""):
    _exec(
        "UPDATE reservations "
        "SET department=%s,name=%s,res_date=%s,res_time=%s,res_time_end=%s,destination=%s,purpose=%s "
        "WHERE id=%s",
        (dept, name, res_date, res_time, res_time_end, dest, purpose, rid),
    )

def delete_reservation(rid):
    _exec("DELETE FROM reservations WHERE id=%s", (rid,))


# ── 운행 기록 ────────────────────────────────────────────────────
def get_pre_drives(user_phone):
    return _query(
        "SELECT * FROM driving_logs WHERE user_phone=%s AND status='pre' "
        "ORDER BY drive_date DESC", (user_phone,))

def get_user_all_logs(user_phone):
    return _query(
        "SELECT * FROM driving_logs WHERE user_phone=%s "
        "ORDER BY drive_date DESC, created_at DESC", (user_phone,))

def get_todays_reservation(user_phone):
    """현재 시간대에 해당하는 오늘 예약을 반환 (목적지·방문목적 자동채우기용)"""
    today = now_kst().strftime("%Y-%m-%d")
    now   = now_kst().strftime("%H:%M")
    rows  = _query(
        "SELECT * FROM reservations WHERE user_phone=%s AND res_date=%s ORDER BY res_time",
        (user_phone, today),
    )
    if not rows:
        return None
    for r in rows:
        if r["res_time"] <= now <= r["res_time_end"]:
            return r
    for r in rows:
        if r["res_time"] >= now:
            return r
    return rows[-1]

def add_pre_drive(user_phone, dept, name, phone,
                  drive_date, odo_start, companions, dest, depart_time="", purpose=""):
    _exec(
        "INSERT INTO driving_logs "
        "(user_phone,department,name,phone,drive_date,"
        "odometer_start,companions,destination,depart_time,purpose) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
        (user_phone, dept, name, phone, drive_date, odo_start, companions, dest, depart_time, purpose),
    )

def complete_drive(lid, odo_end, charge_amt, parking, companions, drive_date, arrive_time="", purpose="", destination=None):
    _exec(
        "UPDATE driving_logs "
        "SET odometer_end=%s,charging_amount=%s,parking_location=%s,"
        "companions=%s,drive_date=%s,arrive_time=%s,purpose=%s,status='complete'"
        + (",destination=%s" if destination is not None else "")
        + " WHERE id=%s",
        (odo_end, charge_amt, parking, companions, drive_date, arrive_time, purpose)
        + ((destination,) if destination is not None else ())
        + (lid,),
    )

def update_drive_log(lid, drive_date, odo_start, odo_end,
                     companions, dest, charge_amt, parking, status,
                     depart_time="", arrive_time="", purpose=""):
    _exec(
        "UPDATE driving_logs "
        "SET drive_date=%s,odometer_start=%s,odometer_end=%s,companions=%s,"
        "destination=%s,charging_amount=%s,parking_location=%s,status=%s,"
        "depart_time=%s,arrive_time=%s,purpose=%s WHERE id=%s",
        (drive_date, odo_start, odo_end, companions,
         dest, charge_amt, parking, status, depart_time, arrive_time, purpose, lid),
    )

def delete_drive_log(lid):
    _exec("DELETE FROM driving_logs WHERE id=%s", (lid,))

def get_logs_by_period(start, end):
    return _query(
        "SELECT * FROM driving_logs WHERE drive_date BETWEEN %s AND %s "
        "AND status='complete' ORDER BY drive_date, created_at",
        (start, end),
    )


# ════════════════════════════════════════════════════════════════
# 엑셀
# ════════════════════════════════════════════════════════════════
def make_excel(logs, start_date, end_date):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "운행기록부"
    ws.sheet_view.showGridLines = False

    thin   = Side(style="thin")
    B_ALL  = Border(left=thin, right=thin, top=thin, bottom=thin)
    CA     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LA     = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    BF     = Font(bold=True, name="맑은 고딕", size=10)
    NF     = Font(name="맑은 고딕", size=10)
    H_FILL = PatternFill("solid", fgColor="BDD7EE")
    G_FILL = PatternFill("solid", fgColor="D9D9D9")

    def mc(r1, c1, r2, c2, val="", font=None, align=CA, border=B_ALL, fill=None):
        if r1 != r2 or c1 != c2:
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        cl = ws.cell(r1, c1, val)
        cl.font, cl.alignment, cl.border = font or NF, align, border
        if fill:
            cl.fill = fill
        return cl

    ws.merge_cells("A1:J1")
    ws["A1"].value, ws["A1"].font = "1. 기본정보", BF
    ws["A1"].fill,  ws["A1"].alignment = G_FILL, LA
    mc(2,1,2,3,"차 종",BF,CA,B_ALL,H_FILL); mc(2,4,2,10,"자동차 등록번호",BF,CA,B_ALL,H_FILL)
    mc(3,1,3,3,VEHICLE_NAME,NF,CA,B_ALL);   mc(3,4,3,10,VEHICLE_NUMBER,NF,CA,B_ALL)
    ws.append([None])
    ws.merge_cells("A5:J5")
    ws["A5"].value, ws["A5"].font = "2. 업무용 사용비율 계산", BF
    ws["A5"].fill,  ws["A5"].alignment = G_FILL, LA
    for r in (6,7,8): ws.row_dimensions[r].height = 24
    mc(6,1,8,1,"사용일자\n(요일)",BF,CA,B_ALL,H_FILL)
    mc(6,2,6,3,"사용자",BF,CA,B_ALL,H_FILL); mc(7,2,8,2,"부서",BF,CA,B_ALL,H_FILL)
    mc(7,3,8,3,"성명",BF,CA,B_ALL,H_FILL);   mc(6,4,6,9,"운 행 내 역",BF,CA,B_ALL,H_FILL)
    mc(7,4,8,4,"주행 전\n계기판의 거리",BF,CA,B_ALL,H_FILL)
    mc(7,5,8,5,"주행 후\n계기판의 거리",BF,CA,B_ALL,H_FILL)
    mc(7,6,8,6,"주행거리\n(km)",BF,CA,B_ALL,H_FILL)
    mc(7,7,7,8,"업무용 사용거리(km)",BF,CA,B_ALL,H_FILL)
    mc(8,7,8,7,"출/퇴근용\n(km)",BF,CA,B_ALL,H_FILL); mc(8,8,8,8,"일반업무용\n(km)",BF,CA,B_ALL,H_FILL)
    mc(7,9,8,9,"비 고",BF,CA,B_ALL,H_FILL);   mc(6,10,8,10,"충전금액",BF,CA,B_ALL,H_FILL)

    dr = 9; td = tc = 0.0
    for log in logs:
        d = datetime.strptime(log["drive_date"], "%Y-%m-%d")
        lbl = f"{d.month:02d}월 {d.day:02d}일({WEEKDAYS[d.weekday()]})"
        s = log["odometer_start"] or 0; e = log["odometer_end"] or 0
        dist = e - s; chrg = log.get("charging_amount") or 0
        td += dist; tc += chrg
        ws.row_dimensions[dr].height = 16
        for col, val in enumerate(
            [lbl, log["department"], log["name"],
             int(s), int(e), int(dist), "X", int(dist),
             log["destination"] or "", int(chrg) if chrg else "-"], 1
        ):
            cl = ws.cell(dr, col, val)
            cl.font, cl.alignment, cl.border = NF, CA, B_ALL
            if isinstance(val, int) and col in (4,5,6,8,10):
                cl.number_format = "#,##0"
        dr += 1

    sr = dr
    ws.row_dimensions[sr].height = 18; ws.row_dimensions[sr+1].height = 22
    mc(sr,1,sr,3,"과세기간 총주행 거리 (km)",BF,CA,B_ALL,H_FILL)
    mc(sr,4,sr,6,int(td),BF,CA,B_ALL,H_FILL); ws.cell(sr,4).number_format="#,##0"
    mc(sr,7,sr,8,"과세기간 업무용 사용거리 (km)",BF,CA,B_ALL,H_FILL)
    mc(sr,9,sr,9,"업무사용비율",BF,CA,B_ALL,H_FILL)
    mc(sr,10,sr,10,"총 충전금액",BF,CA,B_ALL,H_FILL)
    BF2 = Font(bold=True, name="맑은 고딕", size=12)
    mc(sr+1,4,sr+1,6,int(td),BF2,CA,B_ALL); ws.cell(sr+1,4).number_format="#,##0"
    mc(sr+1,7,sr+1,8,int(td),BF2,CA,B_ALL); ws.cell(sr+1,7).number_format="#,##0"
    mc(sr+1,9,sr+1,9,"100%",BF2,CA,B_ALL)
    mc(sr+1,10,sr+1,10,int(tc),BF2,CA,B_ALL); ws.cell(sr+1,10).number_format="#,##0"
    for i, w in enumerate([14,16,8,13,13,10,10,11,20,11],1):
        ws.column_dimensions[get_column_letter(i)].width = w
    out = BytesIO(); wb.save(out); out.seek(0)
    return out


# ════════════════════════════════════════════════════════════════
# 세션 초기화
# ════════════════════════════════════════════════════════════════
def init_session():
    defs = {
        "logged_in":        False,
        "user_phone":       "",
        "user_name":        "",
        "user_department":  DEPARTMENTS[0],
        "user_employee_id": "",
        "admin_logged_in":  False,
        "show_admin_modal": False,
        "selected_cal_date": str(date.today()),
        "date_picker_main": date.today(),   # 달력 클릭 → date_input 자동 반영용
        "cal_year":         date.today().year,
        "cal_month":        date.today().month,
        "adm_logs":       None,
        "editing_log_id": None,
    }
    for k, v in defs.items():
        if k not in st.session_state:
            st.session_state[k] = v



# ════════════════════════════════════════════════════════════════
# 달력 월 네비게이션 — st.button (전체 페이지 로딩 없이 달력만 갱신)
# ════════════════════════════════════════════════════════════════
def render_month_nav(year, month):
    # 마커를 가운데 컬럼 안에 넣어 stHorizontalBlock:has(.mnav-marker)로 직접 타겟
    st.markdown("""
    <style>
    [data-testid="stHorizontalBlock"]:has(.mnav-marker) {
        flex-wrap: nowrap !important;
        gap: 4px !important;
        align-items: center !important;
    }
    [data-testid="stHorizontalBlock"]:has(.mnav-marker) > div:first-child,
    [data-testid="stHorizontalBlock"]:has(.mnav-marker) > div:last-child {
        flex: 0 0 44px !important;
        min-width: 0 !important;
        max-width: 44px !important;
        width: 44px !important;
    }
    [data-testid="stHorizontalBlock"]:has(.mnav-marker) > div:nth-child(2) {
        flex: 1 1 auto !important;
        min-width: 0 !important;
    }
    [data-testid="stHorizontalBlock"]:has(.mnav-marker) button {
        padding: 2px 4px !important;
        min-height: 0 !important;
        min-width: 0 !important;
        width: 100% !important;
        font-size: 1rem !important;
    }
    </style>
    """, unsafe_allow_html=True)

    col_prev, col_label, col_next = st.columns([1, 6, 1])
    with col_prev:
        if st.button("◀", key="btn_cal_prev", use_container_width=True):
            if month == 1:
                st.session_state.cal_year  = year - 1
                st.session_state.cal_month = 12
            else:
                st.session_state.cal_month = month - 1
            st.rerun()
    with col_label:
        st.markdown(
            f"<span class='mnav-marker'></span>"
            f"<div style='text-align:center;font-size:1.3rem;font-weight:700;"
            f"color:#1a3a5c;padding:6px 0'>{year}년 {month}월</div>",
            unsafe_allow_html=True,
        )
    with col_next:
        if st.button("▶", key="btn_cal_next", use_container_width=True):
            if month == 12:
                st.session_state.cal_year  = year + 1
                st.session_state.cal_month = 1
            else:
                st.session_state.cal_month = month + 1
            st.rerun()



def calendar_widget(year, month, all_res, selected_date, today_str):
    """달력 JS 컴포넌트 — 날짜 클릭 시 해당 날짜 문자열 반환"""
    by_date = {}
    for r in all_res:
        by_date.setdefault(r["res_date"], []).append({
            "res_time":     r["res_time"],
            "res_time_end": r.get("res_time_end", ""),
            "name":         r["name"],
        })
    return _cal_component(
        year=year, month=month,
        by_date=by_date,
        selected_date=selected_date,
        today=today_str,
        key="main_calendar",
        default=None,
    )


# ════════════════════════════════════════════════════════════════
# 사이드바
# ════════════════════════════════════════════════════════════════
def render_user_panel():
    st.sidebar.title("👤 사용자")
    if st.session_state.logged_in:
        st.sidebar.success(
            f"**{st.session_state.user_name}**\n\n"
            f"{st.session_state.user_department}\n\n"
            f"📞 {st.session_state.user_phone}"
        )
        if st.sidebar.button("로그아웃", use_container_width=True):
            for k in ["logged_in","user_phone","user_name",
                      "user_department","user_employee_id"]:
                st.session_state[k] = False if k=="logged_in" else ""
            st.rerun()
        return

    tab_l, tab_r = st.sidebar.tabs(["로그인", "계정 등록"])
    with tab_l:
        phone  = st.text_input("전화번호 (ID)", key="si_phone",
                               placeholder="01012345678")
        emp_id = st.text_input("비밀번호", key="si_emp",
                               type="password", placeholder="사번 입력")
        if st.button("로그인", key="btn_login", use_container_width=True):
            if auth_user(phone, emp_id):
                user = get_user(phone)
                st.session_state.update(
                    logged_in=True, user_phone=user["phone"],
                    user_name=user["name"], user_department=user["department"],
                    user_employee_id=user["employee_id"],
                )
                st.rerun()
            else:
                st.error("전화번호 또는 비밀번호가 올바르지 않습니다.")
    with tab_r:
        st.caption("처음 사용 시 한 번만 등록하면 됩니다.")
        r_dept   = st.selectbox("부서", DEPARTMENTS, key="reg_dept")
        r_name   = st.text_input("이름", key="reg_name")
        r_phone  = st.text_input("전화번호 ('-' 없이 입력)", key="reg_phone",
                                 placeholder="01012345678")
        r_emp    = st.text_input("비밀번호 (사번 입력)", key="reg_emp",
                                 type="password")
        r_emp2   = st.text_input("비밀번호 확인", key="reg_emp2",
                                 type="password")
        if st.button("등록 / 정보 수정", key="btn_reg", use_container_width=True):
            if r_name and r_phone and r_emp and r_emp2:
                if r_emp != r_emp2:
                    st.error("비밀번호가 일치하지 않습니다.")
                else:
                    register_user(r_phone, r_emp, r_dept, r_name)
                    st.session_state.update(
                        logged_in=True, user_phone=r_phone, user_name=r_name,
                        user_department=r_dept, user_employee_id=r_emp,
                    )
                    st.success("등록 완료!")
                    st.rerun()
            else:
                st.warning("모든 항목을 입력해 주세요.")


# ════════════════════════════════════════════════════════════════
# 탭 1 : 예약하기
# ════════════════════════════════════════════════════════════════
def _dept_idx(dept):
    return DEPARTMENTS.index(dept) if dept in DEPARTMENTS else 0


def _time_val(t_str, fallback="09:00"):
    try:
        return datetime.strptime(t_str or fallback, "%H:%M").time()
    except ValueError:
        return datetime.strptime(fallback, "%H:%M").time()


def fmt_date(d_str):
    """'2026-05-15' → '2026-05-15 (금)'"""
    try:
        d = datetime.strptime(str(d_str), "%Y-%m-%d")
        return f"{d_str} ({WEEKDAYS[d.weekday()]})"
    except Exception:
        return str(d_str)


# ════════════════════════════════════════════════════════════════
# 팝업 다이얼로그
# ════════════════════════════════════════════════════════════════
@st.dialog("예약 삭제 확인")
def dlg_delete_reservation(r):
    tr = f"{r['res_time']} ~ {r.get('res_time_end','')}"
    st.markdown(f"**날짜:** {r['res_date']}  │  **시간:** {tr}")
    st.write(f"**부서:** {r['department']}  │  **이름:** {r['name']}")
    st.write(f"**방문지:** {r['destination']}")
    if r.get("purpose"):
        st.write(f"**방문 목적:** {r['purpose']}")
    st.warning("이 예약을 삭제하시겠습니까?")
    col1, col2 = st.columns(2)
    if col1.button("삭제 확인", type="primary", use_container_width=True):
        delete_reservation(r["id"])
        st.rerun()
    if col2.button("취소", use_container_width=True):
        st.rerun()


@st.dialog("예약 수정")
def dlg_edit_reservation(r):
    with st.form(f"dlg_edit_res_{r['id']}"):
        e_dept = st.selectbox("부서", DEPARTMENTS, index=_dept_idx(r["department"]))
        e_name = st.text_input("이름", value=r["name"])
        e_date = st.date_input(
            "날짜",
            value=datetime.strptime(r["res_date"], "%Y-%m-%d").date(),
        )
        ec1, ec2 = st.columns(2)
        with ec1:
            e_ts = st.time_input("시작 시간", value=_time_val(r["res_time"]), step=1800)
        with ec2:
            e_te = st.time_input("종료 시간",
                                  value=_time_val(r.get("res_time_end"), "18:00"),
                                  step=1800)
        e_dest    = st.text_input("방문지", value=r["destination"])
        e_purpose = st.text_input("방문 목적", value=r.get("purpose", ""))
        sa, sb = st.columns(2)
        if sa.form_submit_button("저장", type="primary", use_container_width=True):
            if e_name and e_dest:
                conflict = check_reservation_conflict(
                    str(e_date),
                    e_ts.strftime("%H:%M"),
                    e_te.strftime("%H:%M"),
                    exclude_id=r["id"],
                )
                if conflict:
                    st.error("이미 예약된 시간입니다. 다른 시간을 선택해 주세요.")
                else:
                    update_reservation(
                        r["id"], e_dept, e_name, str(e_date),
                        e_ts.strftime("%H:%M"), e_te.strftime("%H:%M"), e_dest, e_purpose,
                    )
                    st.rerun()
            else:
                st.warning("이름과 방문지를 입력해 주세요.")
        if sb.form_submit_button("취소", use_container_width=True):
            st.rerun()


@st.dialog("주행기록 삭제 확인")
def dlg_delete_drive_log(log):
    d = datetime.strptime(log["drive_date"], "%Y-%m-%d")
    st.write(f"**날짜:** {log['drive_date']} ({WEEKDAYS[d.weekday()]})")
    st.write(f"**목적지:** {log['destination'] or '-'}")
    if log["status"] == "complete":
        dist = (log["odometer_end"] or 0) - (log["odometer_start"] or 0)
        st.write(f"**주행 거리:** {dist:,.0f} km")
    st.warning("이 기록을 삭제하시겠습니까?")
    col1, col2 = st.columns(2)
    if col1.button("삭제 확인", type="primary", use_container_width=True):
        delete_drive_log(log["id"])
        st.rerun()
    if col2.button("취소", use_container_width=True):
        st.rerun()


@st.dialog("주행 전 기록 확인")
def dlg_confirm_pre_drive(data):
    st.markdown("#### 아래 내용으로 등록하시겠습니까?")
    st.write(f"**날짜:** {fmt_date(data['date'])}  │  **출발 시간:** {data['depart']}")
    st.write(f"**부서:** {data['dept']}  │  **이름:** {data['name']}")
    st.write(f"**출발 계기판:** {data['odo']:,.0f} km")
    st.write(f"**목적지:** {data['dest']}")
    if data.get("purpose"): st.write(f"**방문 목적:** {data['purpose']}")
    st.write(f"**동행인:** {data['comp'] or '없음'}")
    st.divider()
    col1, col2 = st.columns(2)
    if col1.button("등록", type="primary", use_container_width=True):
        add_pre_drive(
            data["phone"], data["dept"], data["name"], data["phone"],
            data["date"], data["odo"], data["comp"], data["dest"],
            data["depart"], data["purpose"],
        )
        del st.session_state["pending_pre"]
        st.session_state.pop(f"pd_cache_{data['phone']}", None)
        st.rerun()
    if col2.button("취소", use_container_width=True):
        del st.session_state["pending_pre"]
        st.rerun()


@st.dialog("주행 후 기록 확인")
def dlg_confirm_post_drive(data):
    st.markdown("#### 아래 내용으로 완료 처리하시겠습니까?")
    st.write(f"**도착 날짜:** {fmt_date(data['date'])}  │  **도착 시간:** {data['arrive']}")
    st.write(f"**도착 계기판:** {data['odo_end']:,.0f} km  │  **주행 거리:** {data['driven']:,.0f} km")
    if data.get("dest"):    st.write(f"**목적지:** {data['dest']}")
    if data.get("purpose"): st.write(f"**방문 목적:** {data['purpose']}")
    st.write(f"**주차 장소:** {data['park']}")
    st.write(f"**동행인:** {data['comp'] or '없음'}")
    if data.get("charge"):  st.write(f"**충전 금액:** {int(data['charge']):,} 원")
    st.divider()
    col1, col2 = st.columns(2)
    if col1.button("완료 처리", type="primary", use_container_width=True):
        complete_drive(
            data["lid"], data["odo_end"], data["charge"],
            data["park"], data["comp"], data["date"],
            data["arrive"], data["purpose"],
            destination=data.get("dest"),
        )
        del st.session_state["pending_post"]
        # 캐시 초기화 (완료 후 목록 새로고침)
        st.session_state.pop(f"pd_cache_{data.get('phone','')}", None)
        st.session_state.pop(f"pd_odo_s_{data['lid']}", None)
        st.rerun()
    if col2.button("취소", use_container_width=True):
        del st.session_state["pending_post"]
        st.rerun()


def tab_reservation():
    st.subheader("차량 예약 현황")
    all_res = get_all_reservations()

    # ── 달력 월 이동 (항상 한 줄 HTML) ──────────────────────────
    render_month_nav(st.session_state.cal_year, st.session_state.cal_month)

    # ── 달력 컴포넌트 (클릭 → 날짜 자동 반영) ─────────────────────
    clicked = calendar_widget(
        st.session_state.cal_year,
        st.session_state.cal_month,
        all_res,
        st.session_state.selected_cal_date,
        str(date.today()),
    )
    if clicked and clicked != st.session_state.selected_cal_date:
        try:
            d = datetime.strptime(clicked, "%Y-%m-%d").date()
            st.session_state.selected_cal_date = str(d)
            st.session_state.cal_year  = d.year
            st.session_state.cal_month = d.month
            st.session_state["date_picker_main"] = d
            st.rerun()
        except ValueError:
            pass
    st.markdown("---")

    col_left, col_right = st.columns(2)

    # ── 예약 확인 ──────────────────────────────────────────────
    with col_left:
        # 달력 클릭 시 date_picker_main 키로 값이 주입되어 자동 반영
        sel = st.date_input(
            "날짜",
            key="date_picker_main",
            label_visibility="collapsed",
        )
        # 수동으로 날짜를 바꿨을 때 달력 하이라이트도 동기화
        if sel and str(sel) != st.session_state.selected_cal_date:
            st.session_state.selected_cal_date = str(sel)

        sel_res = get_reservations_by_date(str(sel))

        st.markdown(f"##### 📅 {fmt_date(str(sel))} 예약 현황")
        if not sel_res:
            st.info("이 날짜에 예약이 없습니다.")
        else:
            for r in sel_res:
                tr = f"{r['res_time']} ~ {r.get('res_time_end','')}"
                is_mine = (st.session_state.logged_in and
                           r["user_phone"] == st.session_state.user_phone)
                can_edit = is_mine or st.session_state.admin_logged_in
                hdr = f"🕐 {tr}  │  {r['department']} {r['name']}"
                if can_edit: hdr += "  ✏️"

                with st.expander(hdr, expanded=True):
                    st.write(f"**부서:** {r['department']}")
                    st.write(f"**이름:** {r['name']}")
                    st.write(f"**전화번호:** {r['phone']}")
                    st.write(f"**예약 시간:** {tr}")
                    st.write(f"**방문지:** {r['destination']}")
                    if r.get("purpose"):
                        st.write(f"**방문 목적:** {r['purpose']}")

                    if can_edit:
                        ba, bb = st.columns(2)
                        if ba.button("✏️ 수정", key=f"edit_res_btn_{r['id']}", use_container_width=True):
                            dlg_edit_reservation(r)
                        if bb.button("🗑 취소", key=f"del_res_{r['id']}", use_container_width=True):
                            dlg_delete_reservation(r)

    # ── 예약 등록 ──────────────────────────────────────────────
    with col_right:
        st.markdown("#### ✏️ 예약 등록")
        if not st.session_state.logged_in:
            st.warning("예약하려면 왼쪽 사이드바에서 로그인하세요.")
        else:
            with st.form("form_reservation", clear_on_submit=True):
                r1, r2 = st.columns(2)
                with r1:
                    f_dept = st.selectbox("부서", DEPARTMENTS,
                                          index=_dept_idx(st.session_state.user_department))
                with r2:
                    f_name = st.text_input("이름", value=st.session_state.user_name)
                f_date = st.date_input("날짜", value=sel)
                tc1, tc2 = st.columns(2)
                with tc1:
                    f_ts = st.time_input("시작 시간",
                                         value=datetime.strptime("08:00","%H:%M").time(),
                                         step=1800)
                with tc2:
                    f_te = st.time_input("종료 시간",
                                         value=datetime.strptime("17:00","%H:%M").time(),
                                         step=1800)
                d1, d2 = st.columns(2)
                with d1:
                    f_dest = st.text_input("방문지")
                with d2:
                    f_purpose = st.text_input("방문 목적")
                if st.form_submit_button("예약 등록", type="primary",
                                          use_container_width=True):
                    if f_name and f_dest:
                        conflict = check_reservation_conflict(
                            str(f_date),
                            f_ts.strftime("%H:%M"),
                            f_te.strftime("%H:%M"),
                        )
                        if conflict:
                            st.error("이미 예약된 시간입니다. 다른 시간을 선택해 주세요.")
                        else:
                            add_reservation(
                                st.session_state.user_phone,
                                f_dept, f_name, st.session_state.user_phone,
                                str(f_date),
                                f_ts.strftime("%H:%M"), f_te.strftime("%H:%M"),
                                f_dest, f_purpose,
                            )
                            st.success("예약이 등록되었습니다!")
                            st.rerun()
                    else:
                        st.warning("이름과 방문지를 입력해 주세요.")


# ════════════════════════════════════════════════════════════════
# 탭 2 : 주행 전 기록
# ════════════════════════════════════════════════════════════════
def tab_pre_drive():
    st.subheader("주행 전 기록")
    if not st.session_state.logged_in:
        st.warning("사이드바에서 로그인하세요."); return

    # 오늘 예약 자동채우기 — 시간 단위 캐시 (rerun 마다 DB 재조회 방지)
    _res_key = f"today_res_{st.session_state.user_phone}_{now_kst().strftime('%Y-%m-%d-%H')}"
    if _res_key not in st.session_state:
        st.session_state[_res_key] = get_todays_reservation(st.session_state.user_phone)
    matched_res = st.session_state[_res_key]
    if matched_res:
        st.info(
            f"📅 오늘 예약 정보 자동 반영 — "
            f"**{matched_res['res_time']}~{matched_res['res_time_end']}**  │  "
            f"방문지: {matched_res['destination']}"
            + (f"  │  목적: {matched_res['purpose']}" if matched_res.get("purpose") else "")
        )
    auto_dest    = matched_res["destination"] if matched_res else ""
    auto_purpose = matched_res.get("purpose", "") if matched_res else ""

    with st.form("form_pre", clear_on_submit=True):
        r1, r2 = st.columns(2)
        with r1:
            p_dept = st.selectbox("부서", DEPARTMENTS,
                                  index=_dept_idx(st.session_state.user_department))
        with r2:
            p_name = st.text_input("이름", value=st.session_state.user_name)
        p_date = st.date_input("주행 날짜", value=date.today())
        c1, c2 = st.columns(2)
        with c1:
            p_depart = st.time_input("출발 시간",
                                     value=now_kst().replace(second=0, microsecond=0, tzinfo=None).time(),
                                     step=600)
        with c2:
            p_odo = st.number_input("출발 시 계기판 거리 (km)",
                                    min_value=0.0, step=1.0, format="%.0f")
        d1, d2 = st.columns(2)
        with d1:
            p_dest    = st.text_input("목적지", value=auto_dest)
        with d2:
            p_purpose = st.text_input("방문 목적", value=auto_purpose)
        p_comp = st.text_input("동행인", placeholder="부서/이름 형식, 쉼표로 구분")
        if st.form_submit_button("주행 전 기록 저장", type="primary",
                                  use_container_width=True):
            if p_dest and p_odo > 0:
                # 같은 날짜 대기 중인 기록 중복 체크
                _pd_cache_key = f"pd_cache_{st.session_state.user_phone}"
                _existing = st.session_state.get(_pd_cache_key) or get_pre_drives(st.session_state.user_phone)
                if any(p["drive_date"] == str(p_date) for p in _existing):
                    st.warning(f"⚠️ {fmt_date(str(p_date))}에 이미 완료 대기 중인 주행기록이 있습니다. "
                               f"'주행 후 기록' 탭에서 먼저 완료 처리해 주세요.")
                else:
                    st.session_state["pending_pre"] = {
                        "phone": st.session_state.user_phone,
                        "dept": p_dept, "name": p_name,
                        "date": str(p_date), "odo": p_odo,
                        "comp": p_comp, "dest": p_dest,
                        "depart": p_depart.strftime("%H:%M"),
                        "purpose": p_purpose,
                    }
            else:
                st.warning("목적지와 계기판 거리를 입력해 주세요.")

    if "pending_pre" in st.session_state:
        dlg_confirm_pre_drive(st.session_state["pending_pre"])

    _pd_cache_key = f"pd_cache_{st.session_state.user_phone}"
    if _pd_cache_key not in st.session_state:
        st.session_state[_pd_cache_key] = get_pre_drives(st.session_state.user_phone)
    pre = st.session_state[_pd_cache_key]
    if pre:
        st.divider()
        st.markdown("##### 완료 대기 중인 주행 기록")
        for p in pre:
            st.info(
                f"📅 **{fmt_date(p['drive_date'])}**  │  "
                f"출발: **{p['odometer_start']:,.0f} km**  │  "
                f"목적지: {p['destination']}  │  동행: {p['companions'] or '없음'}"
            )


# ════════════════════════════════════════════════════════════════
# 탭 3 : 주행 후 기록
# ════════════════════════════════════════════════════════════════
def tab_post_drive():
    st.subheader("주행 후 기록")
    if not st.session_state.logged_in:
        st.warning("사이드바에서 로그인하세요."); return

    # pre_drives 캐시: odo_end 입력 시 rerun이 발생해도 DB 재조회 방지
    _pd_cache = f"pd_cache_{st.session_state.user_phone}"
    if _pd_cache not in st.session_state:
        st.session_state[_pd_cache] = get_pre_drives(st.session_state.user_phone)
    pre_drives = st.session_state[_pd_cache]

    if not pre_drives:
        st.info("완료 대기 중인 주행 전 기록이 없습니다.\n\n"
                "'주행 전 기록' 탭에서 먼저 기록하세요.")
        return

    options = {
        f"{p['drive_date']}  │  {p['odometer_start']:,.0f} km 출발  │  {p['destination']}": p
        for p in pre_drives
    }
    sel_label = st.selectbox("완료할 주행 기록 선택", list(options.keys()))
    sel_log   = options[sel_label]
    lid       = sel_log["id"]

    # 출발 계기판을 session_state에 저장해 재조회 없이 주행거리 계산에 사용
    _odo_s_key = f"pd_odo_s_{lid}"
    if _odo_s_key not in st.session_state:
        st.session_state[_odo_s_key] = float(sel_log["odometer_start"] or 0)
    odo_s = st.session_state[_odo_s_key]

    with st.container(border=True):
        st.markdown(
            f"**출발 날짜:** {fmt_date(sel_log['drive_date'])}  \n"
            f"**출발 계기판:** {odo_s:,.0f} km  \n"
            f"**목적지:** {sel_log['destination']}  \n"
            f"**동행인:** {sel_log['companions'] or '없음'}"
            + (f"  \n**방문 목적:** {sel_log['purpose']}" if sel_log.get("purpose") else "")
        )

    st.markdown("---")
    with st.form(f"form_post_{lid}"):
        q_date = st.date_input("도착 날짜", value=date.today())
        c1, c2 = st.columns(2)
        with c1:
            q_arrive  = st.time_input("도착 시간",
                                      value=now_kst().replace(second=0, microsecond=0, tzinfo=None).time(),
                                      step=600)
        with c2:
            q_odo_end = st.number_input(
                f"계기판 거리 (km)  [출발: {odo_s:,.0f}]",
                min_value=odo_s, value=odo_s, step=1.0, format="%.0f",
            )
        d1, d2 = st.columns(2)
        with d1:
            q_dest    = st.text_input("목적지", value=sel_log["destination"] or "")
        with d2:
            q_purpose = st.text_input("방문 목적", value=sel_log.get("purpose") or "")
        q_comp       = st.text_input("동행인", value=sel_log["companions"] or "")
        q_park       = st.text_input("주차 장소")
        q_charge_amt = st.number_input(
            "충전 금액 (원, 없으면 0)",
            min_value=0.0, step=100.0, format="%.0f",
        )

        if st.form_submit_button("주행 후 기록 완료", type="primary",
                                  use_container_width=True):
            driven_final = q_odo_end - odo_s
            if not q_park:
                st.warning("주차 장소를 입력해 주세요.")
            elif driven_final < 0:
                st.warning("도착 계기판이 출발 계기판보다 작습니다.")
            else:
                st.session_state["pending_post"] = {
                    "lid": lid, "odo_end": q_odo_end, "charge": q_charge_amt,
                    "park": q_park, "comp": q_comp, "date": str(q_date),
                    "arrive": q_arrive.strftime("%H:%M"), "purpose": q_purpose,
                    "dest": q_dest, "driven": driven_final,
                    "phone": st.session_state.user_phone,
                }

    if "pending_post" in st.session_state:
        dlg_confirm_post_drive(st.session_state["pending_post"])


# ════════════════════════════════════════════════════════════════
# 탭 4 : 내 주행기록
# ════════════════════════════════════════════════════════════════
def _edit_log_form(log):
    lid    = log["id"]
    status = log["status"]
    odo_s  = float(log["odometer_start"] or 0)

    st.markdown("---")
    st.markdown("**주행기록 수정**")
    with st.form(f"form_edit_log_{lid}"):
        ec1, ec2 = st.columns(2)
        with ec1:
            e_date  = st.date_input(
                "날짜",
                value=datetime.strptime(log["drive_date"], "%Y-%m-%d").date(),
            )
            e_odo_s = st.number_input(
                "출발 계기판 (km)", min_value=0.0, value=odo_s,
                step=1.0, format="%.0f",
            )
            e_dest    = st.text_input("목적지", value=log["destination"] or "")
            e_purpose = st.text_input("방문 목적", value=log.get("purpose") or "")
            e_comp    = st.text_input("동행인", value=log["companions"] or "")
        with ec2:
            e_odo_e = None
            if status == "complete":
                e_odo_e = st.number_input(
                    "도착 계기판 (km)", min_value=0.0,
                    value=float(log["odometer_end"] or odo_s),
                    step=1.0, format="%.0f",
                )
                driven = e_odo_e - e_odo_s
                st.metric("주행 거리", f"{driven:,.0f} km")
                e_charge = st.number_input(
                    "충전 금액 (원)",
                    min_value=0.0, value=float(log.get("charging_amount") or 0),
                    step=100.0, format="%.0f",
                )
                e_park = st.text_input("주차 장소",
                                       value=log["parking_location"] or "")
            else:
                st.info("미완료 기록: 출발 정보만 수정 가능합니다.")
                e_charge = 0.0
                e_park   = log["parking_location"] or ""

        sa, sb = st.columns(2)
        if sa.form_submit_button("저장", type="primary", use_container_width=True):
            odo_end = e_odo_e if status == "complete" else None
            update_drive_log(
                lid, str(e_date), e_odo_s, odo_end,
                e_comp, e_dest, e_charge, e_park, status,
                log.get("depart_time") or "", log.get("arrive_time") or "", e_purpose,
            )
            st.session_state.editing_log_id = None
            st.success("수정되었습니다.")
            st.rerun()
        if sb.form_submit_button("취소", use_container_width=True):
            st.session_state.editing_log_id = None
            st.rerun()


def tab_my_logs():
    st.subheader("내 주행기록")
    if not st.session_state.logged_in:
        st.warning("사이드바에서 로그인하세요."); return

    all_logs = get_user_all_logs(st.session_state.user_phone)
    if not all_logs:
        st.info("등록된 주행 기록이 없습니다."); return

    complete = [l for l in all_logs if l["status"] == "complete"]
    pending  = [l for l in all_logs if l["status"] == "pre"]

    total_dist = sum((l["odometer_end"] or 0) - (l["odometer_start"] or 0)
                     for l in complete)
    total_chrg = sum(l.get("charging_amount") or 0 for l in complete)
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("완료된 운행",   f"{len(complete)} 건")
    m2.metric("대기 중인 운행",f"{len(pending)} 건")
    m3.metric("누적 주행거리", f"{total_dist:,.0f} km")
    m4.metric("누적 충전금액", f"{int(total_chrg):,} 원")
    st.divider()

    if complete:
        st.markdown("#### ✅ 완료된 운행 기록")
        for log in complete:
            d    = datetime.strptime(log["drive_date"], "%Y-%m-%d")
            dist = (log["odometer_end"] or 0) - (log["odometer_start"] or 0)
            chrg = log.get("charging_amount") or 0
            hdr  = (f"📅 {log['drive_date']} ({WEEKDAYS[d.weekday()]})  │  "
                    f"{dist:,.0f} km  │  {log['destination'] or '-'}")
            if st.session_state.editing_log_id == log["id"]:
                hdr += "  ✏️"

            with st.expander(hdr):
                ca, cb = st.columns(2)
                with ca:
                    st.write(f"**날짜:** {log['drive_date']} ({WEEKDAYS[d.weekday()]})")
                    if log.get("depart_time"):
                        st.write(f"**출발 시간:** {log['depart_time']}")
                    st.write(f"**출발 계기판:** {log['odometer_start']:,.0f} km")
                    if log.get("arrive_time"):
                        st.write(f"**도착 시간:** {log['arrive_time']}")
                    st.write(f"**도착 계기판:** {log['odometer_end']:,.0f} km")
                    st.write(f"**주행 거리:** {dist:,.0f} km")
                with cb:
                    st.write(f"**목적지:** {log['destination'] or '-'}")
                    if log.get("purpose"):
                        st.write(f"**방문 목적:** {log['purpose']}")
                    st.write(f"**동행인:** {log['companions'] or '없음'}")
                    st.write(f"**주차 장소:** {log['parking_location'] or '-'}")
                    st.write("**충전 금액:** " + (f"{int(chrg):,} 원" if chrg else "-"))

                ba, bb = st.columns(2)
                if ba.button("✏️ 수정", key=f"edit_log_{log['id']}", use_container_width=True):
                    st.session_state.editing_log_id = (
                        None if st.session_state.editing_log_id == log["id"]
                        else log["id"]
                    )
                    st.rerun()
                if bb.button("🗑 삭제", key=f"del_log_{log['id']}", use_container_width=True):
                    dlg_delete_drive_log(log)

                if st.session_state.editing_log_id == log["id"]:
                    _edit_log_form(log)

    if pending:
        st.divider()
        st.markdown("#### ⏳ 주행 완료 대기 중")
        for log in pending:
            d   = datetime.strptime(log["drive_date"], "%Y-%m-%d")
            hdr = (f"📅 {log['drive_date']} ({WEEKDAYS[d.weekday()]})  │  "
                   f"출발 {log['odometer_start']:,.0f} km  │  "
                   f"{log['destination'] or '-'}")
            with st.expander(hdr):
                st.write(f"**날짜:** {log['drive_date']}")
                st.write(f"**출발 계기판:** {log['odometer_start']:,.0f} km")
                st.write(f"**목적지:** {log['destination'] or '-'}")
                st.write(f"**동행인:** {log['companions'] or '없음'}")
                st.caption("'주행 후 기록' 탭에서 완료 처리하세요.")

                ba, bb = st.columns(2)
                if ba.button("✏️ 수정", key=f"edit_pre_{log['id']}", use_container_width=True):
                    st.session_state.editing_log_id = (
                        None if st.session_state.editing_log_id == log["id"]
                        else log["id"]
                    )
                    st.rerun()
                if bb.button("🗑 삭제", key=f"del_pre_{log['id']}", use_container_width=True):
                    dlg_delete_drive_log(log)

                if st.session_state.editing_log_id == log["id"]:
                    _edit_log_form(log)


# ════════════════════════════════════════════════════════════════
# 관리자 화면
# ════════════════════════════════════════════════════════════════
def admin_panel():
    st.subheader("관리자 화면 — 운행기록 조회 및 내보내기")
    c1, c2 = st.columns(2)
    with c1:
        start_d = st.date_input("시작일", value=date.today().replace(day=1),
                                key="adm_start")
    with c2:
        end_d = st.date_input("종료일", value=date.today(), key="adm_end")

    if st.button("조회", type="primary"):
        st.session_state.adm_logs = get_logs_by_period(str(start_d), str(end_d))

    logs = st.session_state.get("adm_logs")
    if logs is None: return
    if not logs:
        st.warning("해당 기간에 완료된 운행 기록이 없습니다."); return

    rows = []
    for log in logs:
        d    = datetime.strptime(log["drive_date"], "%Y-%m-%d")
        dist = (log["odometer_end"] or 0) - (log["odometer_start"] or 0)
        chrg = log.get("charging_amount") or 0
        rows.append({
            "날짜":         f"{log['drive_date']} ({WEEKDAYS[d.weekday()]})",
            "부서":         log["department"],
            "이름":         log["name"],
            "출발시간":     log.get("depart_time") or "-",
            "도착시간":     log.get("arrive_time") or "-",
            "출발 계기판":  f"{log['odometer_start']:,.0f}" if log["odometer_start"] else "-",
            "도착 계기판":  f"{log['odometer_end']:,.0f}"   if log["odometer_end"]   else "-",
            "주행거리(km)": f"{dist:,.0f}",
            "목적지(비고)": log["destination"] or "",
            "방문목적":     log.get("purpose") or "-",
            "충전금액(원)": f"{int(chrg):,}" if chrg else "-",
            "입력시간":     log.get("created_at") or "-",
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    td = sum((l["odometer_end"] or 0)-(l["odometer_start"] or 0) for l in logs)
    tc = sum(l.get("charging_amount") or 0 for l in logs)
    m1, m2, m3 = st.columns(3)
    m1.metric("총 운행 건수", f"{len(logs)} 건")
    m2.metric("총 주행 거리", f"{td:,.0f} km")
    m3.metric("총 충전 금액", f"{int(tc):,} 원")

    st.download_button(
        "📥 엑셀 운행기록부 다운로드",
        data=make_excel(logs, str(start_d), str(end_d)),
        file_name=f"운행기록부_{start_d}_{end_d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )


# ════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════
def main():
    init_db()
    init_session()

    st.markdown("""
    <style>
    /* ── 상단 여백 최소화 ── */
    .block-container,
    [data-testid="stMainBlockContainer"] .block-container {
        padding-top: 2.875rem !important;
    }
    /* ── 탭 바 스크롤 시 상단 고정 ── */
    [data-baseweb="tab-list"] {
        position: sticky !important;
        top: 2.875rem !important;
        z-index: 100 !important;
        background-color: var(--background-color, white) !important;
        box-shadow: 0 2px 6px rgba(0,0,0,0.1) !important;
    }
    @media (prefers-color-scheme: dark) {
        [data-baseweb="tab-list"] {
            background-color: var(--background-color, #0e1117) !important;
        }
    }
    /* ── 달력 컴포넌트 위아래 여백 축소 ── */
    [data-testid="stCustomComponentV1"] {
        margin-top: -0.25rem !important;
        margin-bottom: -0.25rem !important;
    }
    /* ── 버튼 쌍 한 줄 배치 (폼·익스팬더·다이얼로그·border컨테이너) ── */
    [data-testid="stForm"]                       [data-testid="stHorizontalBlock"],
    [data-testid="stExpander"]                   [data-testid="stHorizontalBlock"]:has(button),
    [data-testid="stDialog"]                     [data-testid="stHorizontalBlock"]:has(button),
    [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stHorizontalBlock"]:has(button) {
        flex-wrap: nowrap !important;
    }
    [data-testid="stForm"]                       [data-testid="stHorizontalBlock"]
        > div[data-testid="stColumn"],
    [data-testid="stExpander"]                   [data-testid="stHorizontalBlock"]:has(button)
        > div[data-testid="stColumn"],
    [data-testid="stDialog"]                     [data-testid="stHorizontalBlock"]:has(button)
        > div[data-testid="stColumn"],
    [data-testid="stVerticalBlockBorderWrapper"] [data-testid="stHorizontalBlock"]:has(button)
        > div[data-testid="stColumn"] {
        flex: 1 1 0 !important;
        min-width: 0 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown(
        "<h2 style='margin:0;color:#1a3a5c'>🚗 공용차량 관리 시스템</h2>"
        f"<small style='color:#888'>차종: {VEHICLE_NAME} │ 번호: {VEHICLE_NUMBER}</small>",
        unsafe_allow_html=True,
    )

    render_user_panel()

    if st.session_state.admin_logged_in:
        tabs = st.tabs(["📅 예약하기", "🚀 주행 전 기록",
                        "🏁 주행 후 기록", "📋 내 주행기록", "🔐 관리자 화면"])
        with tabs[0]: tab_reservation()
        with tabs[1]: tab_pre_drive()
        with tabs[2]: tab_post_drive()
        with tabs[3]: tab_my_logs()
        with tabs[4]: admin_panel()
    else:
        tabs = st.tabs(["📅 예약하기", "🚀 주행 전 기록",
                        "🏁 주행 후 기록", "📋 내 주행기록"])
        with tabs[0]: tab_reservation()
        with tabs[1]: tab_pre_drive()
        with tabs[2]: tab_post_drive()
        with tabs[3]: tab_my_logs()

    # ── 관리자 버튼 (페이지 맨 아래) ─────────────────────────────
    _ac1, _ac2 = st.columns([5, 2])
    with _ac2:
        lbl = "🔓 관리자 로그아웃" if st.session_state.admin_logged_in else "🔒 관리자 로그인"
        if st.button(lbl, key="btn_admin", use_container_width=True):
            if st.session_state.admin_logged_in:
                st.session_state.admin_logged_in  = False
                st.session_state.show_admin_modal = False
            else:
                st.session_state.show_admin_modal = True
            st.rerun()

    if st.session_state.show_admin_modal and not st.session_state.admin_logged_in:
        with st.form("adm_auth_form", border=True, clear_on_submit=True):
            st.markdown("#### 관리자 인증")
            pw = st.text_input("비밀번호", type="password")
            b1, b2 = st.columns(2)
            ok     = b1.form_submit_button("확인", type="primary", use_container_width=True)
            cancel = b2.form_submit_button("취소", use_container_width=True)
        if ok:
            if pw == ADMIN_PASSWORD:
                st.session_state.admin_logged_in  = True
                st.session_state.show_admin_modal = False
                st.rerun()
            else:
                st.error("비밀번호가 틀렸습니다.")
        if cancel:
            st.session_state.show_admin_modal = False
            st.rerun()


if __name__ == "__main__":
    main()
